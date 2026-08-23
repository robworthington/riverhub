#!/usr/bin/env python3
"""One-off: build the 'Teign sewage permits to request from the EA' workbook for the colleague.

No Teign permits have been entered into River Hub yet (no config/teign_permits.json, no import),
so the full set of South West Water discharge points inside the Teign catchment bbox is the list to
request. Source = the same EA EDM feeds the catchment importer uses:
  - live outlets (NEH_outlets_PROD) for location, filtered to the Teign bbox
  - the all-years EDM register for the permit reference + site name + asset type, joined by unique id
"""
import json, ssl, sys, urllib.parse, urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE
EDM_ALL_YEARS = ("https://services1.arcgis.com/JZM7qJpmv7vJ0Hzx/arcgis/rest/services/"
                 "edm_annual_returns_all_years_public/FeatureServer/0/query")
REGISTER_URL = "https://environment.data.gov.uk/public-register/view/search-water-discharge-consents"
COMPANY = "South West Water"


def get(url, params):
    with urllib.request.urlopen(f"{url}?{urllib.parse.urlencode(params)}", timeout=120, context=_SSL) as r:
        return json.load(r)


def live_outlets(feed, bbox):
    s, w, n, e = bbox  # lat_min, lon_min, lat_max, lon_max
    d = get(feed, {"where": "1=1", "outFields": "Id,receivingWaterCourse,status,longitude,latitude",
                   "returnGeometry": "false", "f": "json"})
    out = {}
    for f in d["features"]:
        a = f["attributes"]
        lon, lat = a.get("longitude"), a.get("latitude")
        if lon is None or lat is None:
            continue
        if s <= lat <= n and w <= lon <= e:
            out[str(a["Id"]).strip()] = {"rwc": a.get("receivingWaterCourse"), "status": a.get("status"),
                                          "lon": lon, "lat": lat}
    return out


def all_years(company):
    best = {}  # uid -> (year, row)
    offset = 0
    while True:
        d = get(EDM_ALL_YEARS, {
            "where": f"water_company_name = '{company}'",
            "outFields": "unique_id,annual_return_year,storm_discharge_asset_type,"
                         "site_name_wasc_op_name,permit_reference_ea_condat",
            "returnGeometry": "false", "resultOffset": offset, "resultRecordCount": 1000, "f": "json"})
        feats = d.get("features", [])
        for f in feats:
            a = f["attributes"]
            uid = (a.get("unique_id") or "").strip()
            if not uid:
                continue
            try:
                yr = int(str(a.get("annual_return_year")).strip())
            except (TypeError, ValueError):
                yr = 0
            if uid in best and best[uid][0] >= yr:
                continue
            permit = (a.get("permit_reference_ea_condat") or "").strip()
            best[uid] = (yr, {
                "site": (a.get("site_name_wasc_op_name") or "").strip() or None,
                "type": (a.get("storm_discharge_asset_type") or "").strip() or None,
                "permit": None if permit in ("", "#TBC") else permit,
            })
        if len(feats) < 1000:
            break
        offset += len(feats)
    return {uid: v[1] for uid, v in best.items()}


def nice_name(site):
    """'WIDECOMBE_STW_NEWTON ABBOT' -> 'Widecombe STW (Newton Abbot)'-ish; fall back to the raw site."""
    if not site:
        return None
    parts = [p.strip() for p in site.split("_") if p.strip()]
    if len(parts) >= 2:
        name = parts[0].title()
        kind = parts[1].upper()
        town = parts[2].title() if len(parts) >= 3 else None
        return f"{name} {kind}" + (f", {town}" if town else "")
    return site.title()


def read_db_assets(path):
    """TSV from the Teign DB: asset_unique_id<TAB>asset_name<TAB>asset_type<TAB>permit_entered(t/f).
    This is the authoritative catchment-filtered set (point-in-polygon at import)."""
    rows = []
    with open(path) as fh:
        for line in fh:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            uid = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else uid
            atype = parts[2].strip() if len(parts) > 2 else ""
            entered = (parts[3].strip().lower() in ("t", "true", "yes", "1")) if len(parts) > 3 else False
            rows.append({"uid": uid, "db_name": name, "db_type": atype, "entered": entered})
    return rows


TYPE_LABEL = {"sewage_treatment_works": "Treatment works", "combined_sewer_overflow": "CSO",
              "sewage_pumping_station": "Pumping station", "storm_tank": "Storm tank"}


def main():
    assets_path = None
    if "--assets" in sys.argv:
        assets_path = sys.argv[sys.argv.index("--assets") + 1]

    print("-- fetching the EA all-years EDM register for South West Water ...", file=sys.stderr)
    enr = all_years(COMPANY)
    print(f"-- {len(enr)} SWW outlets in the all-years register", file=sys.stderr)

    if assets_path:
        # Authoritative path: real Teign catchment assets from the DB, mapped to permit refs by uid.
        db = read_db_assets(assets_path)
        to_request = [a for a in db if not a["entered"]]
        print(f"-- DB: {len(db)} Teign assets, {len(to_request)} without a permit entered", file=sys.stderr)
        rows = []
        for a in to_request:
            e = enr.get(a["uid"], {})
            rows.append({
                "asset": a["db_name"] or nice_name(e.get("site")) or a["uid"],
                "type": TYPE_LABEL.get(a["db_type"], a["db_type"] or e.get("type")),
                "rwc": None,
                "permit": e.get("permit"),
                "uid": a["uid"],
            })
    else:
        # Fallback: bounding-box superset (includes neighbouring catchments — NOT catchment-accurate).
        bbox = [50.445, -3.975, 50.74, -3.484]  # teign.json geo.bbox
        feed = "https://services-eu1.arcgis.com/OMdMOtfhATJPcHe3/arcgis/rest/services/NEH_outlets_PROD/FeatureServer/0/query"
        print("-- NO --assets file given: falling back to BBOX SUPERSET (over-counts) ...", file=sys.stderr)
        outlets = live_outlets(feed, bbox)
        print(f"-- {len(outlets)} live outlets in the Teign bbox", file=sys.stderr)
        rows = [{"asset": nice_name(enr.get(uid, {}).get("site")) or uid, "type": enr.get(uid, {}).get("type"),
                 "rwc": o.get("rwc"), "permit": enr.get(uid, {}).get("permit"), "uid": uid}
                for uid, o in outlets.items()]

    # treatment works first (carry DWF, most valuable), then by name
    rows.sort(key=lambda r: (0 if "TREAT" in (str(r["type"]) or "").upper() else 1, (r["asset"] or "").lower()))
    with_ref = [r for r in rows if r["permit"]]
    print(f"-- {len(rows)} assets to request, {len(with_ref)} with a permit reference", file=sys.stderr)

    # ---- workbook ----
    HEADERS = ["#", "Asset", "Asset type (EDM)", "Receiving water", "EA permit reference",
               "EDM unique id", "Notes"]
    WIDTHS = [4, 34, 22, 24, 22, 16, 40]
    navy = "1F3864"; lightblue = "D6E4F0"; amber = "FFF2CC"
    thin = Side(style="thin", color="BFBFBF"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = Workbook(); ws = wb.active; ws.title = "Permits to request"
    ncols = len(HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="River Teign — sewage asset permits to request from the Environment Agency")
    c.font = Font(bold=True, size=14, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=navy)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    scope = ("Teign catchment assets in River Hub with no permit entered" if assets_path
             else "BBOX SUPERSET — includes neighbouring catchments; not catchment-accurate")
    sub = ws.cell(row=2, column=1, value=(
        f"{len(rows)} South West Water discharge points ({scope}); {len(with_ref)} carry a permit reference. "
        f"Search the EA public register by the permit reference: {REGISTER_URL}"))
    sub.font = Font(italic=True, size=9, color="555555"); sub.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 42

    HEAD = 4
    for j, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEAD, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); cell.border = border
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS[j - 1]

    for i, r in enumerate(rows):
        rr = HEAD + 1 + i
        is_stw = "TREAT" in (r["type"] or "").upper()
        note = "" if r["permit"] else "No permit reference on EDM — search the register by operator (South West Water) + location."
        vals = [i + 1, r["asset"], r["type"] or "—", r["rwc"] or "—", r["permit"] or "(none on EDM)", r["uid"], note]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=rr, column=j, value=v); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(j in (2, 7)))
        if is_stw:
            for j in range(1, ncols + 1):
                ws.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=amber)
        elif i % 2 == 1:
            for j in range(1, ncols + 1):
                ws.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=lightblue)

    ws.freeze_panes = ws.cell(row=HEAD + 1, column=1)
    ws.print_title_rows = f"{HEAD}:{HEAD}"
    out = "Teign_missing_permits.xlsx"
    wb.save(out)
    print(f"wrote {out}: {len(rows)} assets ({len(with_ref)} with a permit ref)")


if __name__ == "__main__":
    main()
