#!/usr/bin/env python3
"""
Import WINEP (Water Industry National Environment Programme) actions for a catchment — the
EA-issued, legally-binding actions a water company must complete in a price-review cycle.
See ../WINEP-DATA-RESEARCH.md.

PR24/AMP8 (2025–30) is published as a national, geocoded ArcGIS FeatureServer (Rivers Trust mirror
of EA data) carrying per-action driver code, deadline (completionDate), receptor links and
current→proposed permit values. This importer mirrors import_edm.py: one paginated query by water
company + catchment bbox, decode the driver code via the EA DriverCodes lookup, then link each
action to our data and scope to the catchment in SQL —
  * water body  — EA wbID  → water_bodies.ea_water_body_id            (reliable; primary)
  * works/system — normalised works name → sewage_systems.name        (works-level)
  * asset       — nearest sewage_asset within MATCH_DIST_M of the point (outlet-level, where close)
A fetched row is KEPT only if it resolves to one of the above — that drops neighbouring-catchment
rows the rectangular bbox over-captures. Idempotent: replaces the org's rows for this cycle.

PR19 is an XLSX (not geocoded) — handled separately by --pr19-xlsx (see below).

Config-driven (org, water company, bbox). Service URLs pinned here (central connectors), overridable.

Usage:
    python3 import_winep.py [--config config/catchments/<x>.json] > /tmp/winep.sql
    docker run --rm -i postgres:16 psql "$DB_URL" < /tmp/winep.sql

    # PR19 delivery baseline (downloads the national XLSX, links by name/wbID only):
    python3 import_winep.py --pr19-xlsx /path/to/WINEP_PR19.xlsx > /tmp/winep_pr19.sql
"""
import json, os, ssl, sys, urllib.parse, urllib.request
from datetime import datetime, timezone

import catchment_config

_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE

# PR24 geocoded layer (Rivers Trust mirror of EA data) + EA driver-code lookup. Central connectors.
FS = os.environ.get(
    "WINEP_FS_URL",
    "https://services3.arcgis.com/Bb8lfThdhugyc4G3/arcgis/rest/services/"
    "PR24_Water_Industry_National_Environment_Programme/FeatureServer/0",
)
DRIVER_FS = os.environ.get(
    "WINEP_DRIVER_FS_URL",
    "https://services1.arcgis.com/JZM7qJpmv7vJ0Hzx/arcgis/rest/services/DriverCodes/FeatureServer/0",
)
MATCH_DIST_M = 150  # nearest-asset proximity for an outlet-level asset_id (EA loc is ~1km fuzzed)

OUT_FIELDS = ",".join([
    "actionID", "actionComponent", "waterCo", "driverCodePrimary", "driverCodeSecondary",
    "driverCodeTertiary", "tier1Outcome", "actionName", "actionDescription",
    "optionsAssessmentOutcome", "actionCategorisationAim", "spatialScaleActionDelivery",
    "wbID", "wbType", "boundaryName", "completionDate",
    "primarySSSI", "primarySACSPARamsar", "MCZ", "primaryBathingWater", "primaryShellfishWater",
    "currentPermitDWF", "proposedPermitDWF", "currentBODpermit", "proposedBODpermit",
    "currentNH3permit", "proposedNH3permit", "currentPPermit", "proposedPAveragePermit",
])


def get(url, params):
    with urllib.request.urlopen(f"{url}?{urllib.parse.urlencode(params)}", timeout=120, context=_SSL) as r:
        return json.load(r)


def driver_map():
    """Full EA DriverCodes lookup → {code: (obligation, label)}. Small table (~93 rows)."""
    out = {}
    page = get(f"{DRIVER_FS}/query", {
        "where": "1=1", "outFields": "Obligation,Full_Driver_Code,Description_",
        "returnGeometry": "false", "resultRecordCount": 500, "f": "json",
    })
    for f in page.get("features", []):
        a = f["attributes"]
        code = (a.get("Full_Driver_Code") or "").strip()
        if not code:
            continue
        oblig = " ".join((a.get("Obligation") or "").split()) or None
        label = " ".join((a.get("Description_") or "").split()) or None
        out[code] = (oblig, label)
    return out


def fetch(cfg):
    """All actions for the water company (paginated) — NOT bbox-filtered. As with import_edm, the
    catchment can extend past the rectangular bbox (e.g. Princetown on the upper Dart sits west of
    the Dart bbox edge), so a bbox query would wrongly drop edge actions. The SQL scope step below
    (water body / works / asset match) restricts the whole-company set to the catchment. SWW is
    ~782 rows — a handful of pages."""
    company = cfg["company"]["name"]
    feats, offset = [], 0
    while True:
        page = get(f"{FS}/query", {
            "where": f"waterCo LIKE '%{company}%'",
            "outFields": OUT_FIELDS, "returnGeometry": "true", "outSR": 4326,
            "resultOffset": offset, "resultRecordCount": 1000, "f": "json",
        })
        batch = page.get("features", [])
        feats += batch
        if len(batch) < 1000:
            return feats
        offset += len(batch)


def ms_to_date(ms):
    try:
        ms = float(ms)
    except (TypeError, ValueError):
        return None
    if ms <= 0:
        return None
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date().isoformat()


def works_key(action_name):
    """Normalise a WINEP actionName to a bare works key matching sewage_systems.name's prefix
    (system names are '<WORKS> system'). Returns None for free-text/catchment-level action names."""
    if not action_name:
        return None
    s = action_name.upper().replace("\n", " ").strip()
    s = s.replace("SEWAGETREATMENT WORKS", "STW").replace("SEWAGE TREATMENT WORKS", "STW")
    # NAME_TYPE_TOWN convention: the works is the segment before the first underscore
    if "_" in s:
        s = s.split("_")[0].strip()
    # drop parentheticals e.g. "LADY MEADOW STW (NK_WIDECOMBE)"
    if "(" in s:
        s = s.split("(")[0].strip()
    # strip the trailing works-type token(s)
    for suf in (" STW", " ST", " WTW", " WRW", " PS", " SPS", " SEWAGE WORKS"):
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
    # reject obvious free-text (no clear single works) — keep only tidy short keys
    if not s or len(s) > 40 or any(t in s for t in ("OBSTRUCTION", "ABSTRACTION", "RIVER ", "ECOLOGICAL", " RES", "MOOR")):
        return None
    return s.lower()


def q(v):
    return "null" if v is None or v == "" else "'" + str(v).replace("'", "''") + "'"


def attr(a, *names):
    for nm in names:
        v = a.get(nm)
        if v not in (None, ""):
            return " ".join(str(v).split()) if isinstance(v, str) else v
    return None


def main():
    cfg = catchment_config.load()
    org = cfg["org_id"]
    if "--pr19-xlsx" in sys.argv:
        return main_pr19(cfg, org, sys.argv[sys.argv.index("--pr19-xlsx") + 1])

    dmap = driver_map()
    feats = fetch(cfg)
    print(f"-- {len(feats)} WINEP PR24 rows for {cfg['company']['name']} "
          f"({len(dmap)} driver codes decoded)", file=sys.stderr)

    rows = []
    for f in feats:
        a = f["attributes"]; g = f.get("geometry") or {}
        action_id = (a.get("actionID") or "").strip()
        if not action_id:
            continue
        comp = (a.get("actionComponent") or "").strip()
        dcode = (a.get("driverCodePrimary") or "").strip() or None
        oblig, label = dmap.get(dcode, (None, None))
        name = attr(a, "actionName")
        rows.append({
            "cycle": "PR24", "action_id": action_id, "comp": comp,
            "company": attr(a, "waterCo"),
            "dcode": dcode, "dlabel": label, "doblig": oblig,
            "dsec": (a.get("driverCodeSecondary") or "").strip() or None,
            "dter": (a.get("driverCodeTertiary") or "").strip() or None,
            "name": name, "desc": attr(a, "actionDescription"),
            "tier1": attr(a, "tier1Outcome"), "opt": attr(a, "optionsAssessmentOutcome"),
            "aim": attr(a, "actionCategorisationAim"), "scale": attr(a, "spatialScaleActionDelivery"),
            "ea_wb": (a.get("wbID") or "").strip() or None, "wb_type": attr(a, "wbType"),
            "wb_name": attr(a, "boundaryName"), "due": ms_to_date(a.get("completionDate")),
            "bw": attr(a, "primaryBathingWater"), "sw": attr(a, "primaryShellfishWater"),
            "sssi": attr(a, "primarySSSI"), "sac": attr(a, "primarySACSPARamsar"), "mcz": attr(a, "MCZ"),
            "cdwf": attr(a, "currentPermitDWF"), "pdwf": attr(a, "proposedPermitDWF"),
            "cbod": attr(a, "currentBODpermit"), "pbod": attr(a, "proposedBODpermit"),
            "cnh3": attr(a, "currentNH3permit"), "pnh3": attr(a, "proposedNH3permit"),
            "cp": attr(a, "currentPPermit"), "pp": attr(a, "proposedPAveragePermit"),
            "wkey": works_key(name),
            "lon": g.get("x"), "lat": g.get("y"), "source": "winep_pr24_fs",
        })

    if not rows:
        sys.exit("no WINEP rows returned — check the catchment bbox / water company name")
    emit(rows, org, cfg, "PR24")


def emit(rows, org, cfg, cycle):
    orgl = q(org) + "::uuid"
    cols = ("cycle text, action_id text, comp text, company text, dcode text, dlabel text, "
            "doblig text, dsec text, dter text, name text, descr text, tier1 text, opt text, "
            "aim text, scale text, ea_wb text, wb_type text, wb_name text, due date, bw text, "
            "sw text, sssi text, sac text, mcz text, cdwf text, pdwf text, cbod text, pbod text, "
            "cnh3 text, pnh3 text, cp text, pp text, wkey text, lon float, lat float, source text")
    out = [f"-- River Hub: WINEP {cycle} actions for {cfg['river']}. Idempotent (replaces this org's {cycle} rows).",
           "begin;",
           f"create temp table _w({cols}) on commit drop;"]
    for r in rows:
        out.append("insert into _w values (" + ",".join([
            q(cycle), q(r["action_id"]), q(r["comp"]), q(r["company"]), q(r["dcode"]), q(r["dlabel"]),
            q(r["doblig"]), q(r["dsec"]), q(r["dter"]), q(r["name"]), q(r["desc"]), q(r["tier1"]),
            q(r["opt"]), q(r["aim"]), q(r["scale"]), q(r["ea_wb"]), q(r["wb_type"]), q(r["wb_name"]),
            q(r["due"]), q(r["bw"]), q(r["sw"]), q(r["sssi"]), q(r["sac"]), q(r["mcz"]), q(r["cdwf"]),
            q(r["pdwf"]), q(r["cbod"]), q(r["pbod"]), q(r["cnh3"]), q(r["pnh3"]), q(r["cp"]), q(r["pp"]),
            q(r["wkey"]),
            "null" if r["lon"] is None else repr(r["lon"]),
            "null" if r["lat"] is None else repr(r["lat"]), q(r["source"]),
        ]) + ");")

    # resolve our ids, then keep only rows that scope to the catchment (water body OR works OR asset)
    out.append(f"""create temp table _wm on commit drop as
select w.*,
  (select wb.id from water_bodies wb
     where wb.organisation_id = {orgl} and wb.ea_water_body_id = w.ea_wb limit 1) as water_body_id,
  (select sy.id from sewage_systems sy
     where sy.organisation_id = {orgl} and w.wkey is not null
       and lower(split_part(sy.name, ' system', 1)) = w.wkey limit 1) as system_id,
  (select sa.id from sewage_assets sa
     where sa.organisation_id = {orgl} and sa.latitude is not null and sa.longitude is not null
       and w.lon is not null
       and ST_DWithin(ST_SetSRID(ST_MakePoint(sa.longitude, sa.latitude),4326)::geography,
                      ST_SetSRID(ST_MakePoint(w.lon, w.lat),4326)::geography, {MATCH_DIST_M})
     order by ST_Distance(ST_SetSRID(ST_MakePoint(sa.longitude, sa.latitude),4326)::geography,
                          ST_SetSRID(ST_MakePoint(w.lon, w.lat),4326)::geography) limit 1) as asset_id
from _w w;""")

    out.append(f"delete from winep_actions where organisation_id = {orgl} and cycle = {q(cycle)};")
    out.append(f"""insert into winep_actions
  (organisation_id, cycle, action_id, action_component, water_company, driver_code, driver_label,
   driver_obligation, driver_code_secondary, driver_code_tertiary, action_name, action_description,
   tier1_outcome, options_outcome, aim, spatial_scale, ea_water_body_id, wb_type, wb_name,
   water_body_id, asset_id, sewage_system_id, completion_date, bathing_water, shellfish_water, sssi,
   sac_spa_ramsar, mcz, current_permit_dwf, proposed_permit_dwf, current_bod, proposed_bod,
   current_nh3, proposed_nh3, current_p, proposed_p, latitude, longitude, source)
  select distinct on (cycle, action_id, coalesce(comp,''))
    {orgl}, cycle, action_id, coalesce(comp,''), company, dcode, dlabel, doblig, dsec, dter, name, descr, tier1,
    opt, aim, scale, ea_wb, wb_type, wb_name, water_body_id, asset_id, system_id, due, bw, sw, sssi,
    sac, mcz, cdwf, pdwf, cbod, pbod, cnh3, pnh3, cp, pp, lat, lon, source
  from _wm
  where water_body_id is not null or system_id is not null or asset_id is not null
  order by cycle, action_id, coalesce(comp,'');""")
    out.append("select 'WINEP rows fetched: ' || count(*) from _wm;")
    out.append("select 'kept (in catchment): ' || count(*) from _wm "
               "where water_body_id is not null or system_id is not null or asset_id is not null;")
    out.append(f"select 'linked to a works: ' || count(*) from winep_actions "
               f"where organisation_id = {orgl} and cycle = {q(cycle)} and sewage_system_id is not null;")
    out.append(f"select 'winep_actions now ({cycle}): ' || count(*) from winep_actions "
               f"where organisation_id = {orgl} and cycle = {q(cycle)};")
    out.append("commit;")
    print("\n".join(out))


def _norm_hdr(h):
    return " ".join(str(h or "").replace("\n", " ").split()).lower()


def main_pr19(cfg, org, path):
    """PR19/AMP7 delivery baseline. The national WINEP PR19 download is an XLSX (not geocoded — it
    carries OS Eastings/Northings, which we don't convert in v1), so we link by water-body id /
    works name only. Data is on the 'WINEP DATA' sheet under a multi-row title block, with ragged
    header names (trailing spaces, embedded newlines) — so we locate the header row by 'WINEPID' and
    keyword-match columns."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["WINEP DATA"] if "WINEP DATA" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    idx, header = None, []
    for row in rows_iter:
        norm = [_norm_hdr(c) for c in row]
        if "winepid" in norm:               # the header row
            header = norm
            idx = {h: i for i, h in enumerate(norm) if h}
            break
    if idx is None:
        sys.exit("could not locate the WINEP DATA header row (no 'WINEPID' column)")

    def col(row, *keyword_sets):
        """First cell whose header contains all keywords in any set."""
        for kws in keyword_sets:
            for h, i in idx.items():
                if all(k in h for k in kws) and i < len(row) and row[i] not in (None, ""):
                    return " ".join(str(row[i]).split())
        return None

    def parse_date(v):
        if v is None or v == "":
            return None
        if hasattr(v, "isoformat"):         # already a datetime
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
        s = str(v).strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s.split()[0], fmt).date().isoformat()
            except ValueError:
                continue
        return None

    dmap = driver_map()
    company = cfg["company"]["name"].split(" Ltd")[0]
    rows = []
    for row in rows_iter:
        comp = col(row, ["water", "company"])
        if not comp or company.lower() not in comp.lower():
            continue
        name = col(row, ["scheme", "name"], ["name", "investigation"])
        dcode = col(row, ["driver", "code", "primary"], ["driver", "code"])
        oblig, label = dmap.get(dcode, (None, None))
        rows.append({
            "cycle": "PR19",
            "action_id": col(row, ["winepid"], ["unique", "id"]) or name or "",
            "comp": "", "company": comp,
            "dcode": dcode, "dlabel": label, "doblig": oblig,
            "dsec": col(row, ["driver", "code", "secondary"]),
            "dter": col(row, ["driver", "code", "tertiary"]),
            "name": name, "desc": col(row, ["action", "type"], ["measure", "type"]),
            "tier1": None, "opt": col(row, ["certainty"]), "aim": col(row, ["environmental", "outcome"]),
            "scale": None,
            "ea_wb": col(row, ["waterbody", "id"], ["water", "body", "id"]),
            "wb_type": col(row, ["water", "body", "type"]),
            "wb_name": col(row, ["name", "waterbody"], ["name", "water", "body"]),
            "due": parse_date(col_raw(row, idx, ["completion", "date"])),
            "bw": col(row, ["bathing"]), "sw": col(row, ["shellfish"]),
            "sssi": col(row, ["sssi"]), "sac": col(row, ["ramsar"]), "mcz": col(row, ["mcz"]),
            "cdwf": None, "pdwf": None, "cbod": None, "pbod": None, "cnh3": None, "pnh3": None,
            "cp": None, "pp": None, "wkey": works_key(name),
            "lon": None, "lat": None, "source": "winep_pr19_xlsx",
        })
    print(f"-- {len(rows)} WINEP PR19 rows for {company} (XLSX; name/wbID linkage, no geometry)", file=sys.stderr)
    if not rows:
        sys.exit("no PR19 rows for this company — check the XLSX / 'Water Company' column")
    emit(rows, org, cfg, "PR19")


def col_raw(row, idx, kws):
    for h, i in idx.items():
        if all(k in h for k in kws) and i < len(row):
            return row[i]
    return None


if __name__ == "__main__":
    main()
