#!/usr/bin/env python3
"""
Import SWW EDM per-spill start/stop workbooks into spill_events (granular event-level history).
Catchment-agnostic: the 2020/2024/2025/2026 "sww-YYYY-edm-start-stop" workbooks share one layout with
a clean `Unique ID` (SBB…) column == sewage_assets.asset_unique_id, so rows match assets by that id
for either org — no per-catchment crosswalk needed.

We fetch the catchment's outlet IDs from the live feed (by bbox) to filter the (SWW-wide) files down
to this catchment before emitting SQL; the final INSERT still joins sewage_assets by asset_unique_id,
so only real catchment assets load. Idempotent on (asset_id, event_start). Source tag 'sww-edm-history'.

Usage:
  python3 import_edm_startstop.py --config ../config/catchments/dart.json \
      ~/Downloads/sww-2020-edm-start-stops---storm-overflows.xlsx \
      ~/Downloads/sww-2024-edm-start-stop---storm-overflows.xlsx ... > /tmp/dart_startstop.sql
  docker run --rm -i -e PGPASSWORD=.. postgres:16 psql -h .. -U .. -d postgres < /tmp/dart_startstop.sql
"""
import json, os, ssl, sys, urllib.parse, urllib.request
from openpyxl import load_workbook

import catchment_config

_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE


def q(v):
    return "null" if v is None or v == "" else "'" + str(v).replace("'", "''") + "'"


def feed_ids(cfg):
    """SBB outlet ids within the catchment bbox, from the live SWW feed."""
    feed = cfg["company"]["edm_feed"]
    s, w, n, e = cfg["geo"]["bbox"]
    url = feed + "?" + urllib.parse.urlencode({"where": "1=1", "outFields": "Id,longitude,latitude", "returnGeometry": "false", "f": "json"})
    with urllib.request.urlopen(url, timeout=120, context=_SSL) as r:
        feats = json.load(r).get("features", [])
    ids = set()
    for f in feats:
        a = f["attributes"]; lon, lat = a.get("longitude"), a.get("latitude")
        if lon is None or lat is None:
            continue
        if s <= lat <= n and w <= lon <= e:
            ids.add(str(a.get("Id")).strip())
    return ids


def combine(d, t):
    if d is None:
        return None
    dd = d.date() if hasattr(d, "date") else d
    if t is None:
        ts = "00:00:00"
    elif hasattr(t, "strftime"):
        ts = t.strftime("%H:%M:%S")
    else:
        ts = str(t)
    return f"{dd} {ts}"


def read_file(path, ids):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    col = {h: i for i, h in enumerate(header)}
    ci = col.get("Unique ID"); csd = col.get("Discharge Start Date"); cst = col.get("Discharge Start Time")
    ced = col.get("Discharge Stop Date"); cet = col.get("Discharge Stop Time")
    out = []
    for row in it:
        uid = str(row[ci]).strip() if row[ci] is not None else ""
        if uid not in ids:
            continue
        start = combine(row[csd], row[cst])
        if not start:
            continue
        end = combine(row[ced], row[cet]) if ced is not None else None
        out.append((uid, start, end))
    wb.close()
    return out


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--") and a != catchment_config_arg()]
    cfg = catchment_config.load()
    org = cfg["org_id"]; orgl = q(org) + "::uuid"
    ids = feed_ids(cfg)
    print(f"-- {len(ids)} outlet ids in the {cfg['river']} bbox", file=sys.stderr)

    rows = []
    for path in args:
        recs = read_file(path, ids)
        print(f"-- {os.path.basename(path)}: {len(recs)} events matched to the catchment", file=sys.stderr)
        rows += recs
    if not rows:
        sys.exit("no matching events — check the file paths / catchment bbox")
    # de-dup on (uid, start)
    seen = {}
    for uid, s, e in rows:
        seen[(uid, s)] = e
    print(f"-- {len(seen)} distinct (outlet, start) events across {len(args)} files", file=sys.stderr)

    out = [f"-- River Hub: SWW EDM start/stop events for {cfg['river']} (granular history). Idempotent.",
           "begin;",
           "create temp table _ss(unique_id text, event_start timestamptz, event_end timestamptz) on commit drop;"]
    items = list(seen.items())
    for i in range(0, len(items), 1000):
        vals = []
        for (uid, s), e in items[i:i + 1000]:
            vals.append(f"({q(uid)},'{s}'::timestamptz,{'null' if not e else chr(39)+e+chr(39)+'::timestamptz'})")
        out.append("insert into _ss values " + ",".join(vals) + ";")
    out.append(f"""insert into spill_events (organisation_id, asset_id, outlet_id, event_start, event_end, source)
  select {orgl}, a.id, x.unique_id, x.event_start, x.event_end, 'sww-edm-history'
  from _ss x join sewage_assets a on a.organisation_id = {orgl} and a.asset_unique_id = x.unique_id
  on conflict (asset_id, event_start) do update set event_end = excluded.event_end, source = excluded.source;""")
    out.append(f"select 'spill_events total: ' || count(*) from spill_events where organisation_id = {orgl};")
    out.append("commit;")
    print("\n".join(out))


def catchment_config_arg():
    # value following --config (so it isn't treated as a file path)
    a = sys.argv
    for i, v in enumerate(a):
        if v == "--config" and i + 1 < len(a):
            return a[i + 1]
    return None


if __name__ == "__main__":
    main()
