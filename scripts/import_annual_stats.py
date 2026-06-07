#!/usr/bin/env python3
"""
Backfill multi-year EDM spill stats from the EA Storm-Overflow Annual Returns
into edm_annual_stats (see ../CATCHMENT-METHOD.md). One row per outlet per year,
linked to sewage_assets by Unique ID. Columns are matched by HEADER NAME (the
layout drifts between years), not by index.

Usage:
    python3 import_annual_stats.py > /tmp/annual.sql
    docker exec -i supabase_db_river-hub psql -U postgres -d postgres < /tmp/annual.sql   # local
    docker run --rm -i postgres:16 psql "$DB_URL" < /tmp/annual.sql                       # prod
"""
import json, os, ssl, urllib.parse, urllib.request, io, zipfile, datetime
import openpyxl

_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE
CACHE_DIR = os.environ.get("EDM_CACHE", "/tmp")   # expects edm_<year>.zip if pre-downloaded

ORG = "00000000-0000-0000-0000-000000000001"
# 2020 uses a different per-company layout (no WaSC operational name / no Unique ID)
# so it can't be joined reliably — documented gap. 2021-2023 join by operational
# Site Name; 2024 has the SBB Unique ID.
YEARS = [2021, 2022, 2023, 2024]
FILE_DS = "c55e170e-3c75-49a5-8026-a961ff94c8e0"
URL = "https://environment.data.gov.uk/api/file/download"
COMPANY_SHEET_MATCH = "South West Water"


def download_xlsx(year):
    cached = os.path.join(CACHE_DIR, f"edm_{year}.zip")
    if os.path.exists(cached) and os.path.getsize(cached) > 100000:
        data = open(cached, "rb").read()
    else:
        fn = f"EDM_{year}_Storm_Overflow_Annual_Return.zip"
        q = urllib.parse.urlencode({"fileDataSetId": FILE_DS, "fileName": fn})
        with urllib.request.urlopen(f"{URL}?{q}", timeout=180, context=_SSL) as r:
            data = r.read()
    zf = zipfile.ZipFile(io.BytesIO(data))
    name = next(n for n in zf.namelist() if n.endswith(".xlsx") and "all water" in n.lower())
    return io.BytesIO(zf.read(name))


def find_col(headers, *needles, exclude=None):
    for i, h in enumerate(headers):
        hl = str(h or "").lower().replace("\n", " ")
        if all(n.lower() in hl for n in needles) and (not exclude or exclude.lower() not in hl):
            return i
    return None


def dur_to_hours(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime.timedelta):
        return round(v.total_seconds() / 3600, 2)
    if isinstance(v, datetime.time):
        return round(v.hour + v.minute / 60 + v.second / 3600, 2)
    if isinstance(v, (int, float)):
        return round(float(v), 2)                 # already hours (some sheets)
    s = str(v).strip()
    days = 0.0
    if "day" in s:                                 # e.g. "4 days, 12:34:56"
        d, _, rest = s.partition(",")
        try:
            days = float(d.split()[0])
        except (ValueError, IndexError):
            days = 0.0
        s = rest.strip()
    if ":" in s:                                   # hh:mm:ss (hours can exceed 24)
        try:
            parts = [float(x) for x in s.split(":")]
        except ValueError:
            return None
        while len(parts) < 3:
            parts.append(0)
        h, m, sec = parts[0], parts[1], parts[2]
        return round(days * 24 + h + m / 60 + sec / 3600, 2)
    try:
        return round(days * 24 + float(s), 2)
    except ValueError:
        return None


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def q(v):
    return "null" if v is None else "'" + str(v).replace("'", "''") + "'"


def main():
    rows = []   # (outlet_id, year, spill_count, duration_hours, reporting_pct, site)
    for year in YEARS:
        try:
            buf = download_xlsx(year)
        except Exception as e:
            print(f"-- WARN {year}: download failed: {e}")
            continue
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        sheet = next((s for s in wb.sheetnames if COMPANY_SHEET_MATCH.lower() in s.lower()), None)
        if not sheet:
            print(f"-- WARN {year}: no {COMPANY_SHEET_MATCH} sheet")
            continue
        ws = wb[sheet]
        rowvals = list(ws.iter_rows(values_only=True))
        # header = first row containing a "site name" cell
        hidx = next((i for i, r in enumerate(rowvals)
                     if any("site name" in str(c or "").lower() for c in r)), None)
        if hidx is None:
            print(f"-- WARN {year}: no header row")
            continue
        headers = rowvals[hidx]
        c_uid = next((i for i, h in enumerate(headers)
                      if str(h or "").strip().lower() == "unique id"), None)
        c_site = (find_col(headers, "site name", "operational")
                  or find_col(headers, "site name", "wasc")
                  or find_col(headers, "site name"))
        c_dur = find_col(headers, "total duration")
        c_cnt = find_col(headers, "counted spills") or find_col(headers, "spill count")
        c_pct = find_col(headers, "% of reporting") or find_col(headers, "operation", "%")
        n = 0
        for r in rowvals[hidx + 1:]:
            uid = str(r[c_uid]).strip() if c_uid is not None and r[c_uid] else None
            site = str(r[c_site]).strip() if c_site is not None and r[c_site] else None
            if not uid and not site:
                continue
            rows.append((uid, site, year,
                         num(r[c_cnt]) if c_cnt is not None else None,
                         dur_to_hours(r[c_dur]) if c_dur is not None else None,
                         num(r[c_pct]) if c_pct is not None else None))
            n += 1
        print(f"-- {year}: {n} rows (uid={c_uid} site={c_site} cnt={c_cnt} dur={c_dur} pct={c_pct})")

    # Emit SQL: stage all rows, then insert only those matching our assets by
    # Unique ID (2024) OR WaSC operational Site Name (2021-2023).
    print("begin;")
    print("create temp table _annual(uid text, site text, year int, spill_count numeric, total_duration_hours numeric, reporting_pct numeric) on commit drop;")
    for (uid, site, yr, cnt, dur, pct) in rows:
        print(f"insert into _annual values ({q(uid)},{q(site)},{yr},"
              f"{cnt if cnt is not None else 'null'},{dur if dur is not None else 'null'},"
              f"{pct if pct is not None else 'null'});")
    org = q(ORG) + "::uuid"
    print(f"""insert into edm_annual_stats
    (organisation_id, asset_id, outlet_id, year, spill_count, total_duration_hours, reporting_pct, site_name)
  select distinct on (sa.id, a.year)
         {org}, sa.id, sa.asset_unique_id, a.year, a.spill_count, a.total_duration_hours, a.reporting_pct, a.site
  from _annual a
  join sewage_assets sa on sa.organisation_id={org}
       and (sa.asset_unique_id = a.uid or sa.asset_name = a.site)
  order by sa.id, a.year, a.uid nulls last
  on conflict (organisation_id, outlet_id, year) do update set
     asset_id = excluded.asset_id,
     spill_count = excluded.spill_count,
     total_duration_hours = excluded.total_duration_hours,
     reporting_pct = excluded.reporting_pct,
     site_name = excluded.site_name;""")
    print("select 'annual rows in catchment: ' || count(*) from edm_annual_stats;")
    print("commit;")


if __name__ == "__main__":
    main()
