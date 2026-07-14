#!/usr/bin/env python3
"""
Import LEGACY SWW EDM start/stop workbooks (2021-2023) into spill_events (granular event history).

These pre-SBB files key events by inconsistent identifiers, none of which is the SBB unique id used by
sewage_assets:
  - 2021: EA-consents outlet name (often "<WORKS> (n)" for multi-outfall works)
  - 2022: WaSC operational site name (== sewage_assets.asset_name) + some EA "<WORKS> (n)" names
  - 2023: old SWW#### unique id + a short EA name

We build a crosswalk from the EA all-years FeatureServer mapping every known identifier
(old SWW id / WaSC name / EA name) -> canonical SBB, COLLISION-AWARE: a key resolving to more than one
SBB is treated as ambiguous and never guessed.

Multi-outfall works appear in the files as "<WORKS> (1)", "<WORKS> (2)" where the EA works name maps
to >1 SBB (e.g. an SO + SSO). We resolve which number is which outfall by EVIDENCE, not order: sum each
label's granular event-hours per year and assign it to the outfall whose EA annual-return duration it
best matches (accepted only when the fit is unambiguous). This is the same principle as the annual
fix — never conflate co-located outfalls by guessing.

Idempotent: the emitted SQL first deletes the org's existing 2021-2023 sww-edm-history events, then
reloads. Only catchment outlets (live-feed bbox) are kept.

Usage:
  python3 import_edm_startstop_legacy.py --config ../config/catchments/dart.json \
      ~/Downloads/"South West 2021 copy.xlsx" \
      ~/Downloads/"2022 data South West Water copy.xlsx" \
      ~/Downloads/"South West Water 2023 copy.xlsx" > /tmp/dart_legacy.sql
  # add --report for coverage + outfall-assignment audit on stderr, no SQL emitted.
"""
import datetime, json, os, re, ssl, sys, urllib.parse, urllib.request
from openpyxl import load_workbook

import catchment_config
from import_edm_startstop import feed_ids

_SSL = ssl.create_default_context(); _SSL.check_hostname = False; _SSL.verify_mode = ssl.CERT_NONE
FS = os.environ.get(
    "EDM_FS_URL",
    "https://services1.arcgis.com/JZM7qJpmv7vJ0Hzx/arcgis/rest/services/edm_annual_returns_all_years_public/FeatureServer/0",
)
OUT_FIELDS = ",".join([
    "annual_return_year", "unique_id", "old_unique_id_pre_2024", "site_name_wasc_op_name",
    "site_name_ea_condat", "total_spill_duration_hrs_calculated",
])
_SUFFIX = re.compile(r"^(.*?)\s*\((\d+)\)\s*$")  # "BUCKFASTLEIGH STW (2)" -> base, "2"


def q(v):
    return "null" if v is None or v == "" else "'" + str(v).replace("'", "''") + "'"


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().upper()) if s not in (None, "") else ""


def get(url, params):
    with urllib.request.urlopen(f"{url}?{urllib.parse.urlencode(params)}", timeout=120, context=_SSL) as r:
        return json.load(r)


def build_crosswalk(company):
    """Company-wide id/name -> SBB (collision-aware), plus per-(sbb,year) annual duration and the set
    of EA works names that resolve to multiple outfalls (for the (n)-suffix disambiguation)."""
    feats, offset = [], 0
    while True:
        page = get(f"{FS}/query", {
            "where": f"water_company_name = '{company}'", "outFields": OUT_FIELDS,
            "returnGeometry": "false", "resultOffset": offset, "resultRecordCount": 1000, "f": "json",
        })
        batch = page.get("features", [])
        feats += batch
        if len(batch) < 1000:
            break
        offset += len(batch)

    groups = {}  # stable outlet key -> identifiers seen
    for f in feats:
        a = f["attributes"]
        sbb = (a.get("unique_id") or "").strip() or None
        oldid = (a.get("old_unique_id_pre_2024") or "").strip() or None
        wasc = (a.get("site_name_wasc_op_name") or "").strip() or None
        ea = (a.get("site_name_ea_condat") or "").strip() or None
        key = norm(wasc) or norm(ea)
        if not key:
            continue
        g = groups.setdefault(key, {"sbb": set(), "old": set(), "wasc": set(), "ea": set()})
        if sbb: g["sbb"].add(sbb)
        if oldid: g["old"].add(oldid)
        if wasc: g["wasc"].add(norm(wasc))
        if ea: g["ea"].add(norm(ea))

    by_old, by_wasc, by_ea = {}, {}, {}
    key_sbb, sbb_set = {}, set()
    for k, g in groups.items():
        if len(g["sbb"]) != 1:
            continue
        sbb = next(iter(g["sbb"])); sbb_set.add(sbb); key_sbb[k] = sbb
        for o in g["old"]: by_old.setdefault(o, set()).add(sbb)
        for w in g["wasc"]: by_wasc.setdefault(w, set()).add(sbb)
        for e in g["ea"]: by_ea.setdefault(e, set()).add(sbb)

    # per-(sbb,year) EA annual-return duration, sbb resolved via the group's single sbb
    ann_dur = {}
    for f in feats:
        a = f["attributes"]
        sbb = (a.get("unique_id") or "").strip() or key_sbb.get(norm(a.get("site_name_wasc_op_name")) or norm(a.get("site_name_ea_condat")))
        if not sbb:
            continue
        try:
            yr = int(str(a.get("annual_return_year")).strip())
        except (TypeError, ValueError):
            continue
        d = a.get("total_spill_duration_hrs_calculated")
        if d not in (None, ""):
            ann_dur[(sbb, yr)] = float(d)

    uniq = lambda mp: {k: next(iter(v)) for k, v in mp.items() if len(v) == 1}
    ea_multi = {k: sorted(v) for k, v in by_ea.items() if len(v) >= 2}
    return {"sbb": sbb_set, "old": uniq(by_old), "wasc": uniq(by_wasc), "ea": uniq(by_ea),
            "ann_dur": ann_dur, "ea_multi": ea_multi}


def resolve(uid, site, xw):
    if uid:
        u = uid.strip()
        if u in xw["sbb"]: return u
        if u in xw["old"]: return xw["old"][u]
    if site:
        n = norm(site)
        if n in xw["wasc"]: return xw["wasc"][n]
        if n in xw["ea"]: return xw["ea"][n]
        pref = {sbb for k, sbb in xw["ea"].items() if len(n) >= 12 and k.startswith(n)}
        if len(pref) == 1: return next(iter(pref))
    return None


def parse_dt(v):
    if v in (None, ""): return None
    if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d %H:%M:%S")
    s = re.sub(r"\s+", " ", str(v).strip()).replace("T", " ").replace("Z", "")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def hours(start, end):
    if not start or not end: return 0.0
    try:
        a = datetime.datetime.strptime(start, "%Y-%m-%d %H:%M:%S")
        b = datetime.datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
        return max(0.0, (b - a).total_seconds() / 3600)
    except ValueError:
        return 0.0


def find_cols(hdr):
    idx = {}
    for i, h in enumerate(hdr):
        hl = str(h or "").strip().lower()
        if hl in ("unique id", "unique_id"): idx["uid"] = i
        elif hl.startswith("site name") or hl == "site_name": idx.setdefault("site", i)
        elif "discharge start" in hl: idx["start"] = i
        elif "discharge stop" in hl or "discharge end" in hl: idx["end"] = i
    return idx


def read_file(path, xw):
    """Return (resolved events [(sbb,start,end)], ambiguous multi-outfall events
    [(base_norm, label_norm, start, end)], total, unmatched)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    resolved, ambig, total, unmatched = [], [], 0, 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]; it = ws.iter_rows(values_only=True)
        try:
            hdr = list(next(it))
        except StopIteration:
            continue
        idx = find_cols([str(c) if c is not None else "" for c in hdr])
        if "start" not in idx:
            continue
        ci, cs, cst, csite = idx.get("uid"), idx["start"], idx.get("end"), idx.get("site")
        for row in it:
            if cs >= len(row):
                continue
            uid = str(row[ci]).strip() if ci is not None and ci < len(row) and row[ci] is not None else None
            site = str(row[csite]).strip() if csite is not None and csite < len(row) and row[csite] is not None else None
            start = parse_dt(row[cs])
            if not start:
                continue
            total += 1
            end = parse_dt(row[cst]) if cst is not None and cst < len(row) else None
            sbb = resolve(uid, site, xw)
            if sbb is not None:
                resolved.append((sbb, start, end))
                continue
            # "<WORKS> (n)" against an ambiguous EA works name -> defer to duration disambiguation
            m = _SUFFIX.match(site or "")
            if m and norm(m.group(1)) in xw["ea_multi"]:
                ambig.append((norm(m.group(1)), norm(site), start, end))
            else:
                unmatched += 1
        break
    wb.close()
    return resolved, ambig, total, unmatched


def disambiguate(ambig, xw, warn):
    """Assign each '<WORKS> (n)' label to one of the works' SBBs by matching summed granular event-hours
    to the EA annual duration. Returns [(sbb,start,end)] for confident assignments only."""
    by_base = {}
    for base, label, s, e in ambig:
        by_base.setdefault(base, {}).setdefault(label, []).append((s, e))
    out = []
    for base, labels in by_base.items():
        sbbs = xw["ea_multi"].get(base, [])
        if len(labels) != len(sbbs):
            warn(f"skip ambiguous works '{base}': {len(labels)} labels vs {len(sbbs)} outfalls")
            continue
        # granular hours per label per year, and annual hours per sbb per year
        lyears = {lab: {} for lab in labels}
        for lab, evs in labels.items():
            for s, e in evs:
                lyears[lab][s[:4]] = lyears[lab].get(s[:4], 0.0) + hours(s, e)
        years = sorted({y for lab in lyears for y in lyears[lab]})
        # try every label->sbb permutation, pick the one with least total |granular-annual| error
        import itertools
        best, best_err = None, None
        for perm in itertools.permutations(sbbs):
            err = 0.0
            for lab, sbb in zip(labels, perm):
                for y in years:
                    ann = xw["ann_dur"].get((sbb, int(y)))
                    if ann is not None:
                        err += abs(lyears[lab].get(y, 0.0) - ann)
            if best_err is None or err < best_err:
                best, best_err = perm, err
        # confidence: best must clearly beat the next-best permutation
        errs = []
        for perm in itertools.permutations(sbbs):
            err = sum(abs(lyears[lab].get(y, 0.0) - (xw["ann_dur"].get((sbb, int(y)) ) or lyears[lab].get(y, 0.0)))
                      for lab, sbb in zip(labels, perm) for y in years)
            errs.append(err)
        errs.sort()
        confident = len(errs) < 2 or errs[0] <= 0.6 * errs[1]
        assign = dict(zip(labels, best))
        if not confident:
            warn(f"skip ambiguous works '{base}': outfall durations too close to assign safely {dict(assign)}")
            continue
        warn(f"assigned '{base}': " + ", ".join(f"{lab.split('(')[-1].rstrip(')')}->{sbb}" for lab, sbb in assign.items()))
        for lab, evs in labels.items():
            for s, e in evs:
                out.append((assign[lab], s, e))
    return out


def main():
    report = "--report" in sys.argv
    files = [a for a in sys.argv[1:] if not a.startswith("--") and a != _cfg_arg()]
    cfg = catchment_config.load()
    org = cfg["org_id"]; orgl = q(org) + "::uuid"
    xw = build_crosswalk(cfg["company"]["name"])
    print(f"-- crosswalk: {len(xw['sbb'])} SBBs, {len(xw['old'])} old-id, {len(xw['wasc'])} WaSC, "
          f"{len(xw['ea'])} EA, {len(xw['ea_multi'])} multi-outfall works", file=sys.stderr)
    ids = feed_ids(cfg)
    print(f"-- {len(ids)} outlet ids in the {cfg['river']} bbox", file=sys.stderr)
    warn = lambda m: print("--   " + m, file=sys.stderr)

    resolved, ambig = [], []
    for path in files:
        r, am, total, un = read_file(path, xw)
        resolved += r; ambig += am
        print(f"-- {os.path.basename(path)}: {total} rows, {un} unmatched, {len(r)} resolved, "
              f"{len(am)} deferred to outfall-matching", file=sys.stderr)
    resolved += disambiguate(ambig, xw, warn)

    seen = {}
    for sbb, s, e in resolved:
        if sbb in ids:
            seen[(sbb, s)] = e
    print(f"-- {len(seen)} distinct catchment events", file=sys.stderr)
    yc = {}
    for (sbb, s) in seen:
        yc.setdefault(s[:4], set()).add(sbb)
    for y in sorted(yc):
        print(f"--   {y}: {sum(1 for k in seen if k[1][:4]==y)} events / {len(yc[y])} assets", file=sys.stderr)

    if report:
        return
    if not seen:
        sys.exit("no matching catchment events")

    out = [f"-- River Hub: LEGACY SWW EDM start/stop events (2021-2023) for {cfg['river']}. Idempotent.",
           "begin;",
           "create temp table _ss(unique_id text, event_start timestamptz, event_end timestamptz) on commit drop;"]
    items = list(seen.items())
    for i in range(0, len(items), 1000):
        vals = [f"({q(sbb)},'{s}'::timestamptz,{'null' if not e else chr(39)+e+chr(39)+'::timestamptz'})"
                for (sbb, s), e in items[i:i + 1000]]
        out.append("insert into _ss values " + ",".join(vals) + ";")
    out.append(f"""delete from spill_events where organisation_id = {orgl} and source = 'sww-edm-history'
  and event_start >= '2021-01-01' and event_start < '2024-01-01';""")
    out.append(f"""insert into spill_events (organisation_id, asset_id, outlet_id, event_start, event_end, source)
  select {orgl}, a.id, x.unique_id, x.event_start, x.event_end, 'sww-edm-history'
  from _ss x join sewage_assets a on a.organisation_id = {orgl} and a.asset_unique_id = x.unique_id
  on conflict (asset_id, event_start) do update set event_end = excluded.event_end, source = excluded.source;""")
    out.append(f"select 'spill_events 2021-2023: ' || count(*) from spill_events where organisation_id = {orgl} and source = 'sww-edm-history' and event_start >= '2021-01-01' and event_start < '2024-01-01';")
    out.append("commit;")
    print("\n".join(out))


def _cfg_arg():
    a = sys.argv
    for i, v in enumerate(a):
        if v == "--config" and i + 1 < len(a):
            return a[i + 1]
    return None


if __name__ == "__main__":
    main()
