#!/usr/bin/env python3
"""One-off: build the 'Dart sewage permits to request from the EA' workbook for the colleague.
Source: PERMITS-MATCH-AND-GAPS.md section 2 (assets matched to Dart with no permit document held)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REGISTER_URL = "https://environment.data.gov.uk/public-register/view/search-water-discharge-consents"

# Asset, type, discharge point(s), EA permit reference, permit category, priority, notes
ROWS = [
    ("Rattery STW", "Treatment works", "SO", "NRASW1494", "Legacy NRA/SWWA consent", "HIGH",
     "Worst dry-weather spiller & WINEP headline asset — prioritise. Legacy consent only; ask SWW/EA for the current consolidated EPR permit."),
    ("Ashprington STW", "Treatment works", "SO + SSO", "NRASW3983", "Legacy NRA/SWWA consent", "Normal", ""),
    ("Broadhempston STW", "Treatment works", "SSO", "NRASW1075", "Legacy NRA/SWWA consent", "Normal", ""),
    ("Harberton STW", "Treatment works", "SO + SSO", "NRASW5295", "Legacy NRA/SWWA consent", "Normal", ""),
    ("Holne STW", "Treatment works", "SO", "SWWA2251", "Legacy NRA/SWWA consent", "Normal", ""),
    ("Kilbury STW, Buckfastleigh", "Treatment works", "SO + SSO", "NRASW5004 / NRASW5003", "Legacy NRA/SWWA consent", "Normal",
     "Two consent references — one per discharge point."),
    ("Staverton STW", "Treatment works", "SSO", "NRASW0257", "Legacy NRA/SWWA consent", "Normal", ""),
    ("Princetown STW", "Treatment works", "SO + SSO", "201064", "Modern EPR permit", "Normal",
     "Modern permit exists on the register; document not yet held."),
    ("Blackbrook North CSO, Princetown", "CSO", "SO", "201856", "Modern EPR permit", "Normal", ""),
    ("Widecombe STW", "Treatment works", "SO", "203911", "Modern EPR permit", "Normal", ""),
    ("New Park Garden CSO, Widecombe", "CSO", "SO", "201693", "Modern EPR permit", "Normal", ""),
    ("Old Woollen Mill CSO, Buckfastleigh", "CSO", "SO", "201803", "Modern EPR permit", "Normal", ""),
    ("St Lukes Church CSO, Buckfastleigh", "CSO", "SO", "201802", "Modern EPR permit", "Normal", ""),
    ("Pear Tree Cross CSO, Ashburton", "CSO", "SO", "201952", "Modern EPR permit", "Normal", ""),
    ("Stonepark Crescent CSO, Ashburton", "CSO", "SO", "202969", "Modern EPR permit", "Normal", ""),
    ("Stoke Gabriel SPS", "Pumping station", "SO", "202852", "Modern EPR permit", "Normal", ""),
    ("Yarrow Bank SPS, Kingswear", "Pumping station", "SO", "200472", "Modern EPR permit", "Normal", ""),
    ("Opp 26 Market Street CSO, Buckfastleigh", "CSO", "SO", "(no permit reference on EDM)", "Unknown", "Normal",
     "No EA permit reference recorded — search the register by operator (South West Water) + location to identify it."),
]

HEADERS = ["#", "Asset", "Asset type", "Discharge point(s)", "EA permit reference", "Permit type", "Priority", "Notes"]
WIDTHS = [4, 32, 16, 16, 22, 24, 9, 60]

wb = Workbook()
ws = wb.active
ws.title = "Permits to request"

navy = "1F3864"; lightblue = "D6E4F0"; amber = "FFF2CC"
title_font = Font(bold=True, size=14, color="FFFFFF")
sub_font = Font(italic=True, size=9, color="555555")
head_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ncols = len(HEADERS)
# Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
c = ws.cell(row=1, column=1, value="River Dart — sewage asset permits to request from the Environment Agency")
c.font = title_font; c.alignment = Alignment(horizontal="left", vertical="center")
c.fill = PatternFill("solid", fgColor=navy)
ws.row_dimensions[1].height = 24
# Subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
ws.cell(row=2, column=1,
        value=("18 Dart assets matched in River Hub with no permit document held (of 45 total; 23 already entered). "
               "Search the EA public register by the permit reference below: " + REGISTER_URL)).font = sub_font
ws.row_dimensions[2].height = 28
ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

HEAD_ROW = 4
for j, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEAD_ROW, column=j, value=h)
    cell.font = head_font; cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(j)].width = WIDTHS[j - 1]

for i, row in enumerate(ROWS):
    r = HEAD_ROW + 1 + i
    values = [i + 1, *row]
    for j, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(j in (2, 8)))
        if values[6] == "HIGH":
            cell.fill = PatternFill("solid", fgColor=amber)
    if i % 2 == 1 and values[6] != "HIGH":
        for j in range(1, ncols + 1):
            ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=lightblue)

ws.freeze_panes = ws.cell(row=HEAD_ROW + 1, column=1)
ws.print_title_rows = f"{HEAD_ROW}:{HEAD_ROW}"

out = "Dart_missing_permits.xlsx"
wb.save(out)
print(f"wrote {out} with {len(ROWS)} assets")
